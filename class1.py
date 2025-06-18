import time
from argparse import Action

from selenium import webdriver
from selenium.common import NoSuchElementException
from selenium.webdriver import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.support.select import Select
import openpyxl

'''
driver=webdriver.Firefox()

driver.get("https://www.facebook.com/")
driver.maximize_window()
driver.find_element(By.ID,'email').send_keys("manojkumhar")
driver.find_element(By.ID,'pass').send_keys("manoj796")
driver.find_element(By.XPATH,"//button[text()='Log in']").click()
time.sleep(2)

driver.get("http://www.durgasoft.com/registration.asp")
driver.maximize_window()
print(driver.title)
try:
  assert "DURGA SOFT" in driver.title
except:
    print("this element is handled")
driver.find_element(By.ID,'name').send_keys("manojkumhar")
driver.find_element(By.ID,'ph_no').send_keys('9783943489')
driver.find_element(By.ID,'email').send_keys("manojprajapat456@gmail.com")
radiobtn1=driver.find_element(By.XPATH,"//input[@value='Employed']")
radiobtn1.click()
print(radiobtn1.is_selected())
print(radiobtn1.is_displayed())
time.sleep(2)
radiobtn2=driver.find_element(By.ID,"SPRING")
radiobtn2.click()
print(radiobtn2.is_selected())
print(radiobtn2.is_displayed())
time.sleep(2)
radiobtn3=driver.find_element(By.ID,"Unix")
radiobtn3.click()
print(radiobtn3.is_selected())
print(radiobtn3.is_displayed())
driver.find_element(By.ID,"date").send_keys("15-05-2024")
driver.find_element(By.ID,"time").send_keys("15:30")
driver.find_element(By.XPATH,"//textarea[@id='desc']").send_keys("i want to join class")
time.sleep(2)
driver.find_element(By.NAME,'Submit').click()
time.sleep(3)
print(driver.current_url)
print(driver.title)
try:
 driver.find_element(By.PARTIAL_LINK_TEXT,"About Us").click()
except NoSuchElementException as e:
    print("Error will be handled by driver")
list=driver.find_elements(By.TAG_NAME,"a")
print(len(list))
for i in list:
    print(i.text)

driver.get("https://demo.automationtesting.in/Alerts.html")
driver.maximize_window()
driver.find_element(By.XPATH,"//button[@onclick='alertbox()']").click()
l=driver.switch_to.alert
l.accept()
driver.find_element(By.XPATH,"//a[text()='Alert with OK & Cancel ']").click()
driver.find_element(By.XPATH,"//button[text()='click the button to display a confirm box ']").click()
l1=driver.switch_to.alert
l1.dismiss()
time.sleep(2)
driver.find_element(By.XPATH,"//a[text()='Alert with Textbox ']").click()
driver.find_element(By.XPATH,"//button[text()='click the button to demonstrate the prompt box ']").click()
l3=driver.switch_to.alert
l3.send_keys("manoj kumhar")
l3.accept()
time.sleep(3)

driver.get("https://demo.automationtesting.in/Register.html")
driver.maximize_window()
driver.find_element(By.XPATH,"//div[@id='msdd']").click()
select=driver.find_elements(By.XPATH,"//ul[@class='ui-autocomplete ui-front ui-menu ui-widget ui-widget-content ui-corner-all']")
try:
 for i in select.options:
    if i.text=="Greek":
        i.click()
    elif i.text=="English":
        i.click()
    else:
        pass
except:
    print("error is handled")
time.sleep(3)
try:
 drp1=Select(driver.find_element(By.XPATH,"//select[@id='Skills']"))
 drp1.select_by_visible_text('C++')
except:
    print("error handed")
time.sleep(2)
try:
 select2=Select(driver.find_element(By.XPATH,"//select[@id='country']"))
 for i in select2.options:
    if i.text=='Malaysia':
        i.click()
    elif i.text=='Japan':
        i.click()
    else:
        pass
except:
    print("Dropdown is handled")
driver.get("https://testautomationpractice.blogspot.com/")
list2=Select(driver.find_element(By.XPATH,"//select[@id='country']"))
for i in list2.options:
    print(i.text)
list2.select_by_visible_text("France")
print(driver.current_url)
time.sleep(2)
driver.back()
time.sleep(2)
driver.forward()
time.sleep(2)
driver.refresh()
time.sleep(2)
frame1=driver.find_element(By.XPATH,"//iframe[@id='frame-one796456169']")
driver.switch_to.frame(frame1)
driver.find_element(By.XPATH,"//input[@id='RESULT_TextField-0']").send_keys("manojkumhar")
#driver.find_element(By.XPATH,"//input[@id='RESULT_RadioButton-1_1']").click()
driver.find_element(By.XPATH,"//input[@id='RESULT_TextField-2']").send_keys(("05/01/1996"))
drpdwn=Select(driver.find_element(By.XPATH,"//select[@id='RESULT_RadioButton-3']"))
drpdwn.select_by_visible_text("Automation Engineer")
for i in drpdwn.options:
    print(i.text)
driver.find_element(By.XPATH,"//input[@id='FSsubmit']").click()
driver.switch_to.default_content()
driver.find_element(By.XPATH,"//button[text()='New Browser Window']").click()
windowhandes=driver.window_handles
driver.switch_to.window(windowhandes[0])
print(driver.title)
print(driver.current_url)
driver.switch_to.window(windowhandes[1])
print(driver.title)
print(driver.current_url)
driver.quit()
driver.get("https://demo.automationtesting.in/Register.html")
interactions=driver.find_element(By.XPATH,"//a[text()='Interactions ']")
dragdrop=driver.find_element(By.XPATH,"//a[text()='Drag and Drop ']")
video=driver.find_element(By.XPATH,"//a[text()='Video']")
youtube=driver.find_element(By.XPATH,"//a[text()='Youtube']")
action=ActionChains(driver)
action.move_to_element(interactions).perform()
time.sleep(2)
action.move_to_element(dragdrop).perform()
action.move_to_element(video).perform()
action.move_to_element(youtube).perform()
youtube.click()
driver.get("https://demo.automationtesting.in/Static.html")
#drag1=driver.find_element(By.XPATH,"//img[@id='mongo']")
#drag2=driver.find_element(By.XPATH,"//img[@id='node']")
drag=driver.find_element(By.XPATH,"//div[@id='dragarea']")
drop=driver.find_element(By.XPATH,"//div[@id='droparea']")
action.drag_and_drop(drag,drop).perform()
time.sleep(2)

driver.get("https://demo.automationtesting.in/Resizable.html")
resize=driver.find_element(By.XPATH,"//div[@class='ui-resizable-handle ui-resizable-se ui-icon ui-icon-gripsmall-diagonal-se']")
action.drag_and_drop_by_offset(resize,15,20).perform()

driver.get("https://testautomationpractice.blogspot.com/")
copy=driver.find_element(By.XPATH,"//button[@ondblclick='myFunction1()']")
action.double_click(copy).perform()

driver.get("https://testautomationpractice.blogspot.com/")
ele1=driver.find_element(By.XPATH,"//a[text()='orange HRM']")
driver.execute_script("arguments[0].scrollIntoView()",ele1)
time.sleep(3)
driver.execute_script("window.scrollBy(0,document.body.scrollHeight)")
driver.find_element(By.XPATH,"//a[text()='Blogger']")
time.sleep(2)

driver.get("https://demo.automationtesting.in/FileDownload.html")
driver.find_element(By.XPATH,"//a[text()='Download']").click()

#driver.find_element(By.XPATH,"//input[@id='RESULT_FileUpload-10']").send_keys("D:\\samplefile.pdf")

workbook=openpyxl.load_workbook("D://X-WHEEL PANDA CHEATS.xlsx")
sheet=workbook.get_sheet_by_name("BG Symbols")
print(sheet.max_row)
print(sheet.max_column)
print(sheet.cell(1,1).value)
print(sheet.cell(1,2).value)
print(sheet.cell(1,3).value)
print(sheet.cell(3,3).value)
rows=sheet.max_row
colm=sheet.max_column
for r in range(1,rows+1):
    for c in range(1,colm+1):
        print(sheet.cell(r,c).value,end=" ")
    print()

workbook=openpyxl.load_workbook("D://X-WHEEL PANDA CHEATS.xlsx")
sheet=workbook.get_sheet_by_name("JP")
sheet.cell(35,1).value="Manoj"
sheet.cell(36,1).value="prahil"
sheet.cell(37,1).value="shimla"
sheet.cell(38,1).value="tony"
print(sheet.max_row)
print(sheet.max_column)
workbook.save("D://X-WHEEL PANDA CHEATS.xlsx")
rw=sheet.max_row
clm=sheet.max_column
for r in range(1,rw+1):
    for c in range(1,clm+1):
        sheet.cell(r,c).value="Jai"
workbook.save("D://X-WHEEL PANDA CHEATS.xlsx")


driver.get("https://www.facebook.com/")
import logging
logging.basicConfig(filename=".//logfile",
   format='%(asctime)s: %(levelname)s: %(message)s',
   datefmt='%m/%d/%Y %I: %M: %S %p',level=logging.DEBUG)
driver.maximize_window()
driver.find_element(By.ID,'email').send_keys("manojkumhar")
driver.find_element(By.ID,'pass').send_keys("manoj796")
driver.find_element(By.XPATH,"//button[text()='Log in']").click()
driver.save_screenshot("D://titlepge.png")
titlepage=driver.title
print(titlepage)
if titlepage=="Facebook – log in or sign up":
    assert True
    driver.save_screenshot("D://titlepge1.png")
else:
    pass
print(titlepage)

driver.get("https://testautomationpractice.blogspot.com/")
driver.maximize_window()

rows= driver.find_elements(By.XPATH,"//table[@name='BookTable']/tbody/tr")
print (len(rows))
x=len(rows)
l1=[]
for i in range(2,x+1):
    a=driver.find_element(By.XPATH,"//table[@name='BookTable']/tbody/tr["+str(i)+"]/td[4]")
    l1.append(a.text)
print(l1)

COL=driver.find_elements(By.XPATH,"//*[@id='HTML1']/div[1]/table/tbody/tr[1]/th")
for i in COL:
    print(i.text, end="   ")
y=len(COL)
print(len(rows))
r=len(rows)
c=len(COL)
print(len(COL))
print("BookName"+"      "+"Author"+"      "+"Subject"+"      "+"Price")
for r in range (2,r+1):
    for c in range(1,c+1):
        val=driver.find_element(By.XPATH,"//*[@id='HTML1']/div[1]/table/tbody/tr["+str(r)+"]/td["+str(c)+"]").text
        print(val,end='    ')
    print()

driver.get("https://testautomationpractice.blogspot.com/")
rows=driver.find_elements(By.XPATH,"//table[@name='BookTable']/tbody/tr")
for i in rows:
    print(i.text)
colm=driver.find_elements(By.XPATH,"//table[@name='BookTable']/tbody/tr/th")
for j in colm:
  print(j.text)

rowdata=driver.find_element(By.XPATH,"//table[@name='BookTable']/tbody/tr[5]/td")
print(rowdata.text)


columns= driver.find_elements(By.XPATH,"//table[@name='BookTable']/tbody/tr[3]/td")
print (len(columns))
for k in columns:
    print(k.text,end=' ')

columns1= driver.find_elements(By.XPATH,"//table[@name='BookTable']/tbody/tr[1]/th[4]/td")
print (len(columns1))
for l in columns1:
    print(l.text,end=' ')


#to print complete columns
rows1= driver.find_elements(By.XPATH,"//table[@name='BookTable']/tbody/tr")
print (len(rows))
x=len(rows1)
l1=[]
for m in range(2,x+1):
    a=driver.find_element(By.XPATH,"//table[@name='BookTable']/tbody/tr["+str(m)+"]/td[4]")
    l1.append(a.text)
print(l1)

driver.get("https://testautomationpractice.blogspot.com/")
rows=driver.find_elements(By.XPATH,"//table[@name='BookTable']/tbody/tr")
print(len(rows))
x=len(rows)
print(x)
l=[]
for i in range(2,x+1):
    a=driver.find_element(By.XPATH,"//table[@name='BookTable']/tbody/tr["+str(i)+"]/td[2]")
    l.append(a.text)
print(l)

colm=driver.find_elements(By.XPATH,"//table[@name='BookTable']/tbody/tr/th")
y=len(colm)
print(y)
l1=[]
for i in range(2,x+1):
    for j in range(1,y+1):
        b=driver.find_element(By.XPATH,"//table[@name='BookTable']/tbody/tr["+str(i)+"]/td["+str(j)+"]")
        l1.append(b.text)
print(l1)

colm1=driver.find_elements(By.XPATH,"//table[@name='BookTable']/tbody/tr[6]/td")
for k in colm1:
    print(k.text) '''

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time

# Step 1: Open browser
driver = webdriver.Chrome()
driver.get("https://www.makemytrip.com/")
driver.maximize_window()
driver.implicitly_wait(3)

# Step 2: Handle popup
time.sleep(3)
driver.find_element(By.TAG_NAME, 'body').click()

# Step 3: Select From city
driver.find_element(By.XPATH,"//input[@data-cy='fromCity']").click()
driver.find_element(By.XPATH, "//input[@placeholder='From']").send_keys("Delhi")
#wait.until(EC.element_to_be_clickable((By.XPATH, "//p[contains(text(),'Delhi, India')]"))).click()
'''
# Step 4: Select To city
driver.find_element(By.XPATH, "//input[@placeholder='To']").send_keys("Mumbai")
wait.until(EC.element_to_be_clickable((By.XPATH, "//p[contains(text(),'Mumbai, India')]"))).click()

# Step 5: Loop through each month
for month in range(12):
    # Open date picker
    driver.find_element(By.XPATH, "//label[@for='departure']").click()

    # Click "Next Month" in calendar
    for _ in range(month):
        next_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//span[@aria-label='Next Month']")))
        next_button.click()
        time.sleep(1)

    # Select first available date
    available_dates = driver.find_elements(By.XPATH, "//div[@class='DayPicker-Day'][not(contains(@class,'disabled'))]")
    if available_dates:
        available_dates[0].click()
    else:
        print(f"No available dates in month {month+1}")
        continue

    # Step 6: Click Search
    wait.until(EC.element_to_be_clickable((By.XPATH, "//a[text()='Search']"))).click()

    # Step 7: Wait for flight results
    wait.until(EC.presence_of_element_located((By.XPATH, "//div[@class='listingCard']")))

    # Step 8: Get all flight prices
    prices = driver.find_elements(By.XPATH, "//div[contains(@class,'priceSection')]//p")
    price_list = []
    for p in prices:
        price_text = p.text.replace("₹", "").replace(",", "").strip()
        if price_text.isdigit():
            price_list.append(int(price_text))

    # Step 9: Find and click cheapest flight
    if price_list:
        cheapest_price = min(price_list)
        cheapest_index = price_list.index(cheapest_price)
        flights = driver.find_elements(By.XPATH, "//div[@class='listingCard']")
        flights[cheapest_index].click()
        print(f"Month {month+1}: Clicked cheapest flight with ₹{cheapest_price}")
    else:
        print(f"No flights found in month {month+1}")

    time.sleep(5)
    driver.back()  # Go back to search page
    time.sleep(3)

# Step 10: Close browser
driver.quit()'''
l=input("Enter any list:")
l1=[]
for i in l:
    if i not in l1:
        l1.append(i)
    else:
        pass
print(' '.join(l1))